import pandas as pd
import pyodbc
import warnings
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
import google.auth
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload

warnings.filterwarnings("ignore")

# Connect to database and open SQL cursor
print('Connecting to database...')
print('')
conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};'
                      'Trusted_Connection=yes;')
cursor = conn.cursor()

# Query Excel Contacts view to start
print('Querying data from Contacts to Excel view...')
sql = """
SELECT
[Contact ID],
[Display Name],
[Last Name],
[First Name],
[Contact Group],
[Organization],
[Sub Organization],
[Organization Type],
[Title],
[E-Mail],
[Address Line 1],
[Address Line 2],
[City],
[State],
[Zip Code],
[Country],
[Phone],
[Member],
[Membership status],
[Membership Level],
[Archived],
[Event registrant],
[Suspended member],
[Event announcements],
[Member emails and newsletters],
[Email delivery disabled],
[Email delivery disabled automatically],
[Receiving emails disabled],
[Creation date],
[Last login date],
[Interested in volunteering with APRA-IL?],
[Member since],
[Renewal due],
[Renewal date last changed],
[Level last changed],
[Member Value],
[Outstanding Balance],
[Number of Payments],
[Last Payment Date],
[Last Payment Amount],
[Last Purchase Item],
[Events Attended],
[Events Attended-In person],
[Events Attended-Virtual],
[Last Event Date],
[Last Event Name],
[Last Event Location],
[Emails Received],
[Emails Opened],
[Emails Clicked],
[Total Link Clicks],
[Last Email Received Date],
[Last Email Opened Date],
[Last Email Click Date],
[Profile Last Updated],
[Profile last updated by]
FROM [APRA-IL].[dbo].[ExcelContacts]

ORDER BY
[Contact Group],
[Last Name],
[First Name]
"""
df = pd.read_sql(sql,conn)

# Open up pandas ExcelWriter, write Contacts data to Excel sheet
print('Writing Contact data to APRA-IL Data.xlsx Excel sheet...')
print('')
df.to_excel(writer,sheet_name='Contacts')


# Query Excel Events view
print('Querying data from Events to Excel view...')
sql = """
SELECT 
[Event ID],
[Event Name],
[Event Type],
[In-person/Virtual],
[Access Level],
[Location],
[Start Date],
[End Date],
[Registration Enabled],
[Registrations Limit],
[Confirmed Registrations Count],
[Event Revenue],
[Contact ID],
[Display Name],
[Organization],
[Status],
[Contact Group],
[Registration Date],
[Registration Fee],
[Paid Sum],
[Registration Type Name],
[Event Tags],
[Conference],
[Webinar],
[Board],
[Networking],
[Social],
[Educational],
[Membership]
FROM [APRA-IL].[dbo].[ExcelEvents]

ORDER BY
[Start Date] DESC,
[Display Name]
"""
df = pd.read_sql(sql,conn)

# Write Events data to Excel sheet
print('Writing Events data to APRA-IL Data.xlsx Excel sheet...')
print('')
df.to_excel(writer,sheet_name='Events')


# Query Excel Emails view
print('Querying data from Emails to Excel view...')
sql = """
SELECT 
[Email ID],
[Subject],
[Email Type],
[Sender ID],
[Sender Name],
[Sending Type],
[Sent Date],
[Recipient Count],
[SuccessfullySentCount],
[ReadCount],
[FailedCount],
[RecipientsThatClickedAnyLinkCount],
[UniqueLinkClickCount],
[Event ID],
[Event Name],
[Contact ID],
[Recipient Name],
[Last Name],
[First Name],
[Contact Group],
[Organization],
[Is Delivered],
[Is Opened],
[Clicked]
FROM [APRA-IL].[dbo].[ExcelEmails]

ORDER BY
[Sent Date] DESC,
[Last Name],
[First Name]
"""
df = pd.read_sql(sql,conn)

# Write Email data to Excel sheet
print('Writing Email data to APRA-IL Data.xlsx Excel sheet...')
print('')
df.to_excel(writer,sheet_name='Email')


# Query Excel Link Clicks view
print('Querying data from Link Clicks to Excel view...')
sql = """
SELECT 
[Contact ID],
[Recipient Name],
[Contact Group],
[Organization],
[Email ID],
[Sent Date],
[Email Subject],
[Link URL],
[Clicked],
[Link Total Clicks Count],
[Link Category],
[Link Sub Category],
[Event Name]
FROM [APRA-IL].[dbo].[ExcelLinkClicks]
WHERE
[Contact Group] IS NOT NULL

ORDER BY
[Contact Group],
[Recipient Name],
[Sent Date] DESC
"""
df = pd.read_sql(sql,conn)

# Write Link Clicks data to Excel sheet
print('Writing Link Clicks data to APRA-IL Data.xlsx Excel sheet...')
print('')
df.to_excel(writer,sheet_name='LinkClicks')


# Query Excel Payments view
print('Querying data from Payments to Excel view...')
sql = """
SELECT 
[Payment ID],
[Payment Amount],
[Payment Date],
[Payment Type],
[Contact ID],
[Contact Name],
[Contact Last Name],
[Contact First Name],
[Contact Group],
[Contact Organization],
[Tender Type],
[Payment Created By],
[Invoice ID],
[Invoice Date],
[Invoice Type],
[Invoice Created By]
FROM [APRA-IL].[dbo].[ExcelPayments]

ORDER BY
[Payment Date] DESC
"""
df = pd.read_sql(sql,conn)

# Write Payments data to Excel sheet
print('Writing Payments data to APRA-IL Data.xlsx Excel sheet...')
print('')
df.to_excel(writer,sheet_name='Payments')
writer.close()

print('All data written to Excel.')
print('')

# Open up workbook with OpenPyxl
print('Opening up APRA-IL Data.xlsx Excel sheet with OpenPyxl...')
book = load_workbook(path)

# Create a currency style
currency_style = NamedStyle(name="currency_style", number_format='"$"#,##0')

# Loop through sheets, AutoFit all columns, delete first row, format currency fields
print('Looping through sheets, formatting columns...')
for sheet in book:
    sheet.delete_cols(1)

    # setting the column width
    for col in range(1, sheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        
        for row in sheet.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > 90:
                        max_length = 90
                    elif len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

        # format as currency
        column_name = sheet.cell(row=1, column=col).value
        if 'Amount' in column_name or 'Value' in column_name or 'Revenue' in column_name or 'Balance' in column_name:
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=col)
                cell.style = currency_style

print('Finished formatting workbook, closing file.')
book.save(path)   
cursor.close()
